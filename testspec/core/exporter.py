"""Excel exporter — converts cases.md to structured .xlsx files."""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from rich.console import Console

console = Console()


@dataclass
class TestCase:
    """Parsed test case from Markdown."""
    case_id: str = ""
    title: str = ""
    module: str = ""
    priority: str = ""
    test_type: str = ""
    preconditions: str = ""
    steps: str = ""
    expected: str = ""
    test_data: str = ""


def parse_cases_markdown(content: str) -> list[TestCase]:
    """Parse cases.md into structured TestCase objects."""
    cases: list[TestCase] = []

    # Split by test case headers (### TC-xxx: title)
    pattern = r"###\s+(TC-[^\n:：]+)[：:]\s*([^\n]+)"
    blocks = re.split(pattern, content)

    # blocks[0] is text before first TC, then triplets of (id, title, body)
    i = 1
    while i + 2 <= len(blocks):
        case_id = blocks[i].strip()
        title = blocks[i + 1].strip()
        body = blocks[i + 2] if i + 2 < len(blocks) else ""
        i += 3

        tc = TestCase(case_id=case_id, title=title)

        # Extract fields from body
        tc.module = _extract_field(body, r"\*\*所属模块\*\*[：:]\s*(.+)")
        tc.priority = _extract_field(body, r"\*\*优先级\*\*[：:]\s*(.+)")
        tc.test_type = _extract_field(body, r"\*\*测试类型\*\*[：:]\s*(.+)")
        tc.preconditions = _extract_block(body, r"\*\*前置条件\*\*[：:]", r"\*\*操作步骤\*\*")
        tc.steps = _extract_block(body, r"\*\*操作步骤\*\*[：:]", r"\*\*预期结果\*\*")
        tc.expected = _extract_block(body, r"\*\*预期结果\*\*[：:]", r"(?:\*\*测试数据\*\*|###|$)")
        tc.test_data = _extract_field(body, r"\*\*测试数据\*\*[：:]\s*(.+)")

        cases.append(tc)

    return cases


def _extract_field(text: str, pattern: str) -> str:
    match = re.search(pattern, text)
    return match.group(1).strip() if match else ""


def _extract_block(text: str, start_pattern: str, end_pattern: str) -> str:
    match = re.search(
        f"{start_pattern}(.*?)(?={end_pattern})",
        text,
        re.DOTALL,
    )
    if not match:
        return ""
    block = match.group(1).strip()
    # Clean up list markers and join
    lines = [
        re.sub(r"^\s*\d+\.\s*", "", line).strip()
        for line in block.splitlines()
        if line.strip()
    ]
    return "\n".join(lines)


def export_excel(
    cases_md_path: Path,
    output_path: Path,
    feature_name: str = "",
) -> Path:
    """Export cases.md to a formatted Excel file."""
    content = cases_md_path.read_text(encoding="utf-8")
    cases = parse_cases_markdown(content)

    if not cases:
        console.print("[yellow]Warning: No test cases parsed from the markdown file.[/yellow]")
        console.print("[dim]Make sure cases.md follows the TC-xxx format.[/dim]")

    wb = Workbook()
    ws = wb.active
    ws.title = feature_name or "测试用例"

    # ── Header style ──
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = [
        "用例编号", "用例标题", "所属模块", "优先级", "测试类型",
        "前置条件", "操作步骤", "预期结果", "测试数据",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # ── Data rows ──
    cell_alignment = Alignment(vertical="top", wrap_text=True)

    # Priority color mapping
    priority_colors = {
        "P0": PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
        "P1": PatternFill(start_color="FFA940", end_color="FFA940", fill_type="solid"),
        "P2": PatternFill(start_color="FFD666", end_color="FFD666", fill_type="solid"),
        "P3": PatternFill(start_color="95DE64", end_color="95DE64", fill_type="solid"),
    }

    for row_idx, tc in enumerate(cases, 2):
        values = [
            tc.case_id, tc.title, tc.module, tc.priority, tc.test_type,
            tc.preconditions, tc.steps, tc.expected, tc.test_data,
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border

        # Color-code priority column
        p_cell = ws.cell(row=row_idx, column=4)
        p_key = tc.priority.strip().upper()[:2]
        if p_key in priority_colors:
            p_cell.fill = priority_colors[p_key]
            p_cell.font = Font(bold=True)

    # ── Column widths ──
    col_widths = [14, 30, 16, 10, 12, 30, 40, 40, 20]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width

    # ── Freeze top row ──
    ws.freeze_panes = "A2"

    # ── Auto-filter ──
    ws.auto_filter.ref = f"A1:I{len(cases) + 1}"

    wb.save(output_path)
    console.print(f"[green]✓[/green] Excel exported → {output_path}")
    console.print(f"[dim]  Total test cases: {len(cases)}[/dim]")
    return output_path
